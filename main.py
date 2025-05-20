import logging.config
import logging
import os.path
import re
from contextlib import asynccontextmanager
from io import BytesIO

from fastapi import FastAPI, Request, UploadFile, File, HTTPException, Form
from starlette.templating import Jinja2Templates
from fastapi.responses import HTMLResponse
from openpyxl.reader.excel import load_workbook

from logging_config import LOGGING_CONFIG
from utilits.csrf import generate_csrf_token, validate_csrf_token
from utilits.data_comporator import compare
from utilits.redis_controller import redis_init


templates = Jinja2Templates(directory="templates")

UPLOAD_DIR = 'uploads'
os.makedirs(UPLOAD_DIR, exist_ok=True)

@asynccontextmanager
async def lifespan(app: FastAPI):
    app.state.redis = await redis_init()

    yield

    await app.state.redis.close()
    logger.info("Redis closed")

app = FastAPI(lifespan=lifespan)

logging.config.dictConfig(LOGGING_CONFIG)
logger = logging.getLogger("app")

@app.get("/")
async def root():
    return {"message": "this is root endpoint from fastApiMic"}

@app.get("/hello/{name}")
async def say_hello(name: str):
    return {"message": f"Hello {name}"}

@app.get("/upload_form/", response_class=HTMLResponse)
async def upload_from(request: Request):
    csrf_token = generate_csrf_token()
    return templates.TemplateResponse("upload_form.html", {"request": request, "csrf_token": csrf_token})

@app.post("/upload/")
async def upload_file(
        csrf_token: str = Form(...),
        file1: UploadFile = File(...),
        file2: UploadFile = File(...)):

    files = [file1, file2]
    if not validate_csrf_token(csrf_token):
        raise HTTPException(status_code=403, detail="CSRF token is invalid")
    result = []
    for file in files:
        try:
            if not file.filename.endswith(".xlsx"):
                raise HTTPException(status_code = 400, detail="xlsx files only")

            buffer = BytesIO()
            while chunk := await file.read(1024 * 1024):
                buffer.write(chunk)

            wb = load_workbook(buffer, data_only=True, read_only=True)
            data = []
            if len(wb.sheetnames) >= 1:
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
            else:
                return None
            try:
                for row in sheet.iter_rows(values_only=True):
                    for idx, cell in enumerate(row):
                        if isinstance(cell, int) or re.search(r'^[0-9]+$', str(cell)):
                            row = row[idx:]
                            data.append(row)
                            break
            except Exception as e:
                logger.error(f'exel parse file failure with error: {e}')

            wb.close()

            compare(data, data)

        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code = 500, detail=str(e))
        finally:
            await file.close()






